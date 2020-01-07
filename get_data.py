import openpyxl as opx
import json
from get_geo import get_coord

PATH = 'json/data.json'


def get_shape_sheet(work_sheet):
    flag_n = False
    rows = work_sheet.max_row
    cols = work_sheet.max_column
    for i in range(1, rows + 1):
        for j in range(1, cols + 1):
            if '№' in str(work_sheet.cell(row=i, column=j).value) and flag_n == False:
                y1 = j
                if not work_sheet.cell(row=i + 1, column=j).value is None:
                    x1 = i + 1
                elif work_sheet.cell(row=i + 1, column=j).value is None and not work_sheet.cell(row=i + 2,
                                                                                                column=j).value is None:
                    x1 = i + 2
                else:
                    print('Ошибка!', work_sheet, 'row', i, 'column', j)
                    return False
                for a in range(i + 2, rows + 2):
                    if work_sheet.cell(row=a, column=j).value is None:
                        x2 = a - 1
                        y2 = cols
                        shape_sheet = (x1, y1, x2, y2)
                        flag_n = True
                        break
    return shape_sheet


def get_data_sheet(shape_sheet, work_sheet):
    x1, y1, x2, y2 = shape_sheet
    data_sheet = {}
    for a in range(x1, x2 + 1):
        adres = 'Севастополь, ' + str(work_sheet.cell(row=a, column=y1 + 2).value) + ', ' + str(
            work_sheet.cell(row=a, column=y1 + 3).value)
        row = {'id': str(work_sheet.cell(row=a, column=y1).value),
               'district': str(work_sheet.cell(row=a, column=y1 + 1).value),
               'street': str(work_sheet.cell(row=a, column=y1 + 2).value),
               'house': str(work_sheet.cell(row=a, column=y1 + 3).value),
               'year_built': str(work_sheet.cell(row=a, column=y1 + 4).value),
               'square': str(work_sheet.cell(row=a, column=y1 + 5).value),
               'date_inc_license': str(work_sheet.cell(row=a, column=y1 + 6).value),
               'coordinates': get_coord(adres)}
        '''for b in range(y1, y2 + 1):
            row[b - 1] = str(work_sheet.cell(row=a, column=b).value)'''
        data_sheet[a - x1] = row
    return data_sheet


def save_data_workbook_json(data):
    with open(PATH, 'w', encoding='utf-8') as write_file:
        json.dump(data, write_file, ensure_ascii=False, indent=4)


work_book = opx.load_workbook(filename='sev.xlsx')
sheet_list = work_book.sheetnames
data_workbook = {}
for i in range(1, len(sheet_list)):
    work_book.active = i
    ws = work_book.active
    sheet_name = str(ws.title).replace('"', '').replace(' ', '_')
    data_workbook[sheet_name] = get_data_sheet(get_shape_sheet(ws), ws)
print(len(data_workbook))
save_data_workbook_json(data_workbook)
